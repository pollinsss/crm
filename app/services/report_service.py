"""
Сервис генерации отчётов (PDF, Excel).
Зависимости: reportlab, openpyxl
"""
import io
import os
from datetime import date
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from app.models.order import Order, OrderStatus, FurnitureType, STATUS_LABELS


def _get_font_path() -> str | None:
    """Ищет TTF-шрифт с кириллицей, при необходимости скачивает"""
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "/usr/local/share/fonts/dejavu/DejaVuSans.ttf",
        "/System/Library/Fonts/DejaVuSans.ttf",
        "/System/Library/Fonts/Arial.ttf",
        "C:\\Windows\\Fonts\\arial.ttf",
        "C:\\Windows\\Fonts\\Arial.ttf",
        "C:\\Windows\\Fonts\\calibri.ttf",
        "C:\\Windows\\Fonts\\times.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            return path

    for search_root in ["/usr/share/fonts", "/usr/local/share/fonts", os.path.expanduser("~/.fonts")]:
        if os.path.exists(search_root):
            for root, dirs, files in os.walk(search_root):
                for f in files:
                    if f.lower() in ("dejavusans.ttf", "arial.ttf"):
                        full = os.path.join(root, f)
                        return full

    font_dir = os.path.join(os.path.dirname(__file__), "fonts")
    os.makedirs(font_dir, exist_ok=True)
    local_path = os.path.join(font_dir, "DejaVuSans.ttf")
    if os.path.exists(local_path):
        return local_path

    try:
        import urllib.request
        url = "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans.ttf"
        urllib.request.urlretrieve(url, local_path)
        return local_path
    except Exception:
        return None


def _register_font():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.fonts import addMapping

    font_path = _get_font_path()
    if font_path and os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont("CyrillicFont", font_path))
        addMapping("CyrillicFont", 0, 0, "CyrillicFont")
        return "CyrillicFont"
    return "Helvetica"


FURNITURE_LABELS: dict[FurnitureType, str] = {
    FurnitureType.KITCHEN: "Кухня",
    FurnitureType.WARDROBE: "Шкаф",
    FurnitureType.BED: "Кровать",
    FurnitureType.TABLE: "Стол",
    FurnitureType.CHAIR: "Стул",
    FurnitureType.SOFA: "Диван",
    FurnitureType.HALLWAY: "Прихожая",
    FurnitureType.OTHER: "Другое",
}


async def fetch_orders_for_period(
    db: AsyncSession, date_from: date, date_to: date
) -> list[Order]:
    result = await db.execute(
        select(Order).where(
            and_(
                Order.created_at >= date_from,
                Order.created_at <= date_to,
            )
        ).order_by(Order.created_at)
    )
    return result.scalars().all()


def generate_pdf_report(orders: list[Order], date_from: date, date_to: date) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    font_name = _register_font()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    styles["Normal"].fontName = font_name
    styles["Heading1"].fontName = font_name

    story = []

    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=14, spaceAfter=6, fontName=font_name)
    story.append(Paragraph(f"Отчёт по заказам: {date_from} — {date_to}", title_style))
    story.append(Spacer(1, 12))

    total_revenue = sum(o.final_price for o in orders if o.status == OrderStatus.COMPLETED)
    completed_count = sum(1 for o in orders if o.status == OrderStatus.COMPLETED)
    story.append(Paragraph(f"Выполненных заказов: {completed_count}", styles["Normal"]))
    story.append(Paragraph(f"Общая выручка: {total_revenue:,.0f} руб.", styles["Normal"]))
    story.append(Spacer(1, 16))

    headers = ["№ заказа", "Клиент", "Тип мебели", "Статус", "Сумма", "Маржа"]
    data = [headers]
    for o in orders:
        furniture_label = FURNITURE_LABELS.get(o.furniture_type, o.furniture_type.value)
        status_label = STATUS_LABELS.get(o.status, o.status.value)
        data.append([
            o.order_number,
            str(o.client_id),
            furniture_label,
            status_label,
            f"{o.final_price:,.0f}",
            f"{o.margin:.1f}%",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D9E75")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1FFF8")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ALIGN", (4, 1), (5, -1), "RIGHT"),
    ]))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def generate_excel_report(orders: list[Order], date_from: date, date_to: date) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"

    header_fill = PatternFill("solid", fgColor="1D9E75")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["№ заказа", "ID клиента", "Менеджер", "Тип мебели", "Статус",
               "Цена", "Себестоимость", "Скидка %", "Итоговая сумма", "Маржа %", "Создан"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, o in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=o.order_number)
        ws.cell(row=row, column=2, value=o.client_id)
        ws.cell(row=row, column=3, value=o.manager_id)
        furniture_label = FURNITURE_LABELS.get(o.furniture_type, o.furniture_type.value)
        status_label = STATUS_LABELS.get(o.status, o.status.value)
        ws.cell(row=row, column=4, value=furniture_label)
        ws.cell(row=row, column=5, value=status_label)
        ws.cell(row=row, column=6, value=o.price)
        ws.cell(row=row, column=7, value=o.cost_price)
        ws.cell(row=row, column=8, value=o.discount)
        ws.cell(row=row, column=9, value=o.final_price)
        ws.cell(row=row, column=10, value=o.margin)
        ws.cell(row=row, column=11, value=str(o.created_at.date()))

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()